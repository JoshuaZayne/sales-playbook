"""
Report Generator
================

Generate Excel reports for sales activities.
"""

from pathlib import Path
from typing import Dict, Any, Optional, List
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ReportGenerator:
    """Generate professional Excel reports for sales activities."""

    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for report generation. Install with: pip install openpyxl")

        # Styles
        self.header_font = Font(bold=True, size=12, color="FFFFFF")
        self.header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        self.title_font = Font(bold=True, size=14)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def generate_discovery_report(
        self,
        prospect_name: str,
        qualification_scores: Optional[Dict] = None,
        notes: Optional[List[str]] = None,
        output_path: str = "discovery_report.xlsx"
    ) -> str:
        """
        Generate a discovery report.

        Args:
            prospect_name: Name of the prospect
            qualification_scores: Optional qualification results
            notes: Optional list of discovery notes
            output_path: Output file path

        Returns:
            Path to generated file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Discovery Summary"

        # Title
        ws['A1'] = f"Discovery Report: {prospect_name}"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:D1')

        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A2'].font = Font(italic=True, size=10)

        # Add discovery questions by stage
        row = 4
        from discovery.question_framework import QuestionFramework
        framework = QuestionFramework()

        for stage in ['initial', 'technical', 'business', 'decision']:
            ws.cell(row=row, column=1, value=f"{stage.upper()} STAGE QUESTIONS")
            ws.cell(row=row, column=1).font = Font(bold=True, size=11)
            ws.cell(row=row, column=1).fill = PatternFill(start_color="D9E2F3", fill_type="solid")
            ws.merge_cells(f'A{row}:D{row}')
            row += 1

            questions = framework.get_questions_by_stage(stage)
            for q in questions:
                ws.cell(row=row, column=1, value=q['question'])
                ws.cell(row=row, column=2, value=q['purpose'])
                ws.cell(row=row, column=3, value="")  # Notes column
                row += 1

            row += 1  # Blank row between stages

        # Adjust column widths
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40

        wb.save(output_path)
        return output_path

    def generate_qualification_report(
        self,
        prospect_name: str,
        bant_scores: Optional[Dict] = None,
        meddic_scores: Optional[Dict] = None,
        output_path: str = "qualification_report.xlsx"
    ) -> str:
        """
        Generate a qualification scorecard report.

        Args:
            prospect_name: Name of the prospect
            bant_scores: BANT scoring results
            meddic_scores: MEDDIC scoring results
            output_path: Output file path

        Returns:
            Path to generated file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Qualification Scorecard"

        # Title
        ws['A1'] = f"Qualification Scorecard: {prospect_name}"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:D1')

        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        # BANT Section
        row = 4
        ws.cell(row=row, column=1, value="BANT FRAMEWORK")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).fill = self.header_fill
        ws.cell(row=row, column=1).font = self.header_font
        ws.merge_cells(f'A{row}:C{row}')
        row += 1

        bant_criteria = [
            ("Budget", "Is budget allocated for this initiative?", "/25"),
            ("Authority", "Are we engaged with the decision maker?", "/25"),
            ("Need", "How urgent is the business need?", "/25"),
            ("Timeline", "When do they need to implement?", "/25")
        ]

        for criterion, description, max_score in bant_criteria:
            ws.cell(row=row, column=1, value=criterion)
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2, value=description)
            ws.cell(row=row, column=3, value=f"__{max_score}")
            row += 1

        ws.cell(row=row, column=1, value="TOTAL BANT SCORE")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=3, value="__/100")
        row += 2

        # MEDDIC Section
        ws.cell(row=row, column=1, value="MEDDIC FRAMEWORK")
        ws.cell(row=row, column=1).font = self.header_font
        ws.cell(row=row, column=1).fill = self.header_fill
        ws.merge_cells(f'A{row}:C{row}')
        row += 1

        meddic_criteria = [
            ("Metrics", "What metrics will measure success?"),
            ("Economic Buyer", "Who controls the budget?"),
            ("Decision Criteria", "What are their evaluation criteria?"),
            ("Decision Process", "What is the approval process?"),
            ("Identify Pain", "What problems are they solving?"),
            ("Champion", "Who is advocating internally?")
        ]

        for criterion, description in meddic_criteria:
            ws.cell(row=row, column=1, value=criterion)
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2, value=description)
            ws.cell(row=row, column=3, value="[ ] Defined  [ ] Partial  [ ] Unknown")
            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 35

        wb.save(output_path)
        return output_path

    def generate_roi_report(
        self,
        prospect_name: str,
        roi_results: Optional[Dict] = None,
        output_path: str = "roi_report.xlsx"
    ) -> str:
        """
        Generate an ROI analysis report.

        Args:
            prospect_name: Name of the prospect
            roi_results: ROI calculation results
            output_path: Output file path

        Returns:
            Path to generated file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "ROI Analysis"

        # Title
        ws['A1'] = f"ROI Analysis: {prospect_name}"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:D1')

        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        # Input section
        row = 4
        ws.cell(row=row, column=1, value="INPUTS")
        ws.cell(row=row, column=1).font = self.header_font
        ws.cell(row=row, column=1).fill = self.header_fill
        ws.merge_cells(f'A{row}:C{row}')
        row += 1

        inputs = [
            ("Current Annual Cost", "$"),
            ("Efficiency Gain", "%"),
            ("Implementation Cost", "$"),
            ("Annual License", "$"),
            ("Analysis Period", "years")
        ]

        for label, unit in inputs:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=f"__________ {unit}")
            row += 1

        row += 1

        # Results section
        ws.cell(row=row, column=1, value="RESULTS")
        ws.cell(row=row, column=1).font = self.header_font
        ws.cell(row=row, column=1).fill = self.header_fill
        ws.merge_cells(f'A{row}:C{row}')
        row += 1

        results = [
            "Annual Savings",
            "Net Annual Benefit",
            "Total Investment",
            "3-Year Net Benefit",
            "ROI %",
            "Payback Period (months)"
        ]

        for label in results:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=1).font = Font(bold=True) if "ROI" in label or "Payback" in label else None
            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20

        wb.save(output_path)
        return output_path

    def generate_full_report(
        self,
        prospect_name: str,
        output_path: str = "full_report.xlsx"
    ) -> str:
        """
        Generate a comprehensive report with all sections.

        Args:
            prospect_name: Name of the prospect
            output_path: Output file path

        Returns:
            Path to generated file
        """
        wb = Workbook()

        # Discovery sheet
        ws1 = wb.active
        ws1.title = "Discovery"
        ws1['A1'] = f"Discovery Notes: {prospect_name}"
        ws1['A1'].font = self.title_font

        # Qualification sheet
        ws2 = wb.create_sheet("Qualification")
        ws2['A1'] = f"Qualification Scorecard: {prospect_name}"
        ws2['A1'].font = self.title_font

        # ROI sheet
        ws3 = wb.create_sheet("ROI Analysis")
        ws3['A1'] = f"ROI Analysis: {prospect_name}"
        ws3['A1'].font = self.title_font

        # Next Steps sheet
        ws4 = wb.create_sheet("Next Steps")
        ws4['A1'] = f"Action Items: {prospect_name}"
        ws4['A1'].font = self.title_font

        wb.save(output_path)
        return output_path
